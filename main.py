import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from pdfplumber.utils import extract_text


class Parser:
    def __init__(self, url):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        self.url = url
        self.html_content = self.get_page(url)
        self.soup = BeautifulSoup(self.html_content, 'html.parser') if self.html_content else None
        self.all_tables = self._extract_all_tables() if self.soup else []
        self.date = self.extract_date()

        self.text = self.soup.get_text().replace("fgos.ru", "").replace(str(self.date), "") if self.soup else ""
        self.doc_name = self.extract_doc_name()

    def get_page(self, url):
        try:
            response = self.session.get(url)
            response.raise_for_status()
            return response.text
        except requests.RequestException as e:
            print(f"Ошибка при получении страницы: {e}")
            return None

    def _extract_all_tables(self):
        all_tables_data = []

        try:
            tables = pd.read_html(str(self.soup))

            for i, table_df in enumerate(tables):
                table_data = table_df.fillna('').values.tolist()

                table_data = [[str(cell) for cell in row] for row in table_data]

                all_tables_data.append({
                    'table_obj': None,
                    'data': table_data,
                    'html': '',
                    'index': i
                })

        except Exception as e:
            print(f"Ошибка при парсинге таблиц pandas: {e}")
            all_tables_data = []
            tables = self.soup.find_all('table')

            for i, table in enumerate(tables):
                table_data = []
                rows = table.find_all('tr')

                for row in rows:
                    cols = row.find_all(['th', 'td'])
                    row_data = [col.get_text(strip=True) for col in cols]
                    if any(row_data):
                        table_data.append(row_data)

                all_tables_data.append({
                    'table_obj': table,
                    'data': table_data,
                    'html': str(table),
                    'index': i
                })

        return all_tables_data

    def extract_date(self):
        date_pattern = r'\d{1,2}\.\d{1,2}\.\d{4}'
        text = self.soup.get_text()

        date_match = re.search(date_pattern, text)
        if date_match:
            return date_match[0]


    def extract_doc_name(self):
        code_pattern = r'\d{2}\.[0][345]\.\d{2}'
        date_pattern = r'от\s*\d{1,2}\s+\w+\s+\d{4}\s*г\.'

        found_code = None
        found_date = None

        code_match = re.search(code_pattern, self.text)
        if code_match and not found_code:
            found_code = code_match.group()

        date_match = re.search(date_pattern, self.text)
        if date_match and not found_date:
            found_date = date_match.group()

        if found_code and found_date:
            doc_name = f"{found_code}_{found_date.replace(' ', '_')}"
        elif found_code:
            doc_name = found_code
        elif found_date:
            doc_name = found_date

        return doc_name

    def extract_discipliny_specialiteta(self):
        pattern1 = r'2\.2\.[\sа-яА-Я\(\):]+по'
        pattern2 = r'В федеральных государственных организациях[\,\(\)а-яА-Я\s\.]+2\.3.'
        pattern3= r'2\.3\.'

        match1 = re.search(pattern1, self.text)
        match2 = re.search(pattern2, self.text)
        if not match2:
            match2 = re.search(pattern3, self.text)


        if match1 and match2:
            start_pos = match1.end()
            end_pos = match2.start()
            disciplines_text = self.text[start_pos:end_pos].strip()
            disciplines_text = disciplines_text.replace(';', ',').replace(".", "")
            disciplines = [discipline.strip() for discipline in disciplines_text.split(',')]
            a = ''
            i = 0
            while i < len(disciplines):
                if "(" in disciplines[i]:
                    a = []
                    n = [i]
                    a.append(disciplines[i])
                    j = i + 1
                    while j < len(disciplines) and ")" not in disciplines[j]:
                        a.append(disciplines[j])
                        n.append(j)
                        j += 1

                    if j < len(disciplines):
                        a.append(disciplines[j])
                        n.append(j)

                    for q in n[::-1]:
                        disciplines.pop(q)

                    for z in range(len(a)):
                        disciplines.insert(n[0] + z, a[z])

                    i = n[0] + 1
                else:
                    i += 1
            for i in range(len(disciplines)):
                if "(" in disciplines[i] and '"' not in disciplines[i]:
                    for j in range(len(disciplines[i])):
                        if disciplines[i][j] == "(":
                            a = j + 1
                            break
                    disciplines[i] = str(disciplines[i][a:])

                if ")" in disciplines[i] and '"' not in disciplines[i]:
                    for j in range(len(disciplines[i])):
                        if disciplines[i][j] == ")":
                            b = j
                            break
                    disciplines[i] = str(disciplines[i][:b])

            list_disciplin = []
            for discipline in disciplines:
                list_disciplin.append(discipline.replace("\n", " "))
            return list_disciplin

    def extract_specializacii(self):
        pattern_direct = r'специализация\s№\s*(\d+)\s*"([^"]+(?:"\s*\([^)]+\))?)"'
        pattern2=r'1.14.[а-яА-я\s]+:'
        pattern3=r'1.15'
        matches = re.findall(pattern_direct, self.text)

        numbers_s = []
        names_s = []

        if matches:
            for match in matches:
                num = match[0]
                name = match[1]
                if num not in numbers_s:
                    numbers_s.append(num)
                    names_s.append(name)
        if not numbers_s and not names_s:
            match1 = re.search(pattern2, self.text)
            match2 = re.search(pattern3, self.text)
            if match1 and match2:
                start_pos = match1.end()
                end_pos = match2.start()
                specializacii_text = self.text[start_pos:end_pos].strip()
                specializacii = [discipline.strip() for discipline in specializacii_text.split(';')]
                for i in range(len(specializacii)):
                    numbers_s.append(i+1)
                    names_s.append((specializacii[i]))

        return numbers_s, names_s



    def extract_structure_and_volume(self):
        structure_keywords = ['структура программы', 'объем программы', 'блок', 'дисциплины', 'практика']

        for table_info in self.all_tables:
            table_text = ' '.join([' '.join(row) for row in table_info['data']]).lower()

            keyword_count = sum(1 for keyword in structure_keywords if keyword in table_text)

            if keyword_count >= 2:

                return table_info['data']


    def extract_praktika(self):
        pattern1 = r'Типы учебной практики:'
        pattern2 = r'Тип[\ы]? производственной практики:'
        pattern3 = r'Преддипломная практика проводится для выполнения выпускной квалификационной работы и является обязательной.'
        pattern4= r'2\.5\.'
        if self.doc_name[4]=="4":
            pattern4=r'2\.3\.'
        match1 = re.search(pattern1, self.text)
        match2 = re.search(pattern2, self.text)
        match3 = re.search(pattern3, self.text)
        if not match3:
            match3 = re.search(pattern4, self.text)
        #
        uch_praktika = []
        pr_praktika = []
        if match1 and match2:
            start_pos = match1.end()
            end_pos = match2.start()
            uch_praktika_text = self.text[start_pos:end_pos].strip()

            sentences = uch_praktika_text.split(';')
            clean_sentences = []
            for sentence in sentences:
                sentence = sentence.strip()
                if sentence:
                    if sentence.endswith('.'):
                        sentence = sentence[:-1]
                    clean_sentences.append(sentence)

            uch_praktika = clean_sentences

        if match2 and match3:
            start_pos = match2.end()
            end_pos = match3.start()
            pr_praktika_text = self.text[start_pos:end_pos].strip()

            sentences = pr_praktika_text.split(';')
            clean_sentences = []
            for sentence in sentences:
                sentence = sentence.strip()
                if sentence:
                    if sentence.endswith('.'):
                        sentence = sentence[:-1]
                    clean_sentences.append(sentence)

            pr_praktika = clean_sentences

        return uch_praktika, pr_praktika

    def extract_uk(self):
        for table_info in self.all_tables:
            table_data = table_info['data']

            table_text = ' '.join([' '.join(row) for row in table_data])
            if 'УК-' in table_text:
                return table_data

        return []


    def extract_opk(self):

        pattern1 = r'3\.3\. Программа [\w\s]+:'
        pattern2 = r'3\.4\. Профессиональные компетенции'

        match1 = re.search(pattern1, self.text)
        match2 = re.search(pattern2, self.text)

        if match1 and match2:
            start_pos = match1.end()
            end_pos = match2.start()
            opk_text = self.text[start_pos:end_pos].strip()

            opk_items = []
            i = 0
            while i < len(opk_text):
                if (i <= len(opk_text) - 3 and
                        opk_text[i] == "О" and
                        opk_text[i + 1] == "П" and
                        opk_text[i + 2] == "К"):

                    next_opk = opk_text.find("ОПК", i + 3)
                    if next_opk == -1:
                        next_opk = len(opk_text)

                    opk_item = opk_text[i:next_opk].strip()
                    semicolon_pos = opk_item.find(';')
                    if semicolon_pos != -1:
                        opk_item = opk_item[:semicolon_pos]

                    opk_items.append(opk_item)
                    i = next_opk
                else:
                    i += 1

            processed_opk = []

            for opk in opk_items:
                if opk:
                    parts = opk.split('.', 1)
                    if len(parts) == 2:
                        code = parts[0].strip()
                        description = parts[1].strip().replace("\n", " ").replace(';', '')

                        if re.match(r'^\d+\.', description):
                            digit_match = re.match(r'^(\d+)\.\s*(.+)', description)
                            if digit_match:
                                digit = digit_match.group(1)
                                clean_description = digit_match.group(2)
                                new_code = f"{code}.{digit}"
                                processed_opk.append([new_code, clean_description])
                            else:
                                processed_opk.append([code, description])
                        else:
                            processed_opk.append([code, description])

            for opk in processed_opk:
                if "." in opk[1]:
                    dot_index = opk[1].find(".")
                    if dot_index != -1:
                        opk[1] = opk[1][:dot_index]

            def opk_key(item):
                code = item[0]
                numbers = re.findall(r'\d+', code)
                return tuple(map(int, numbers))

            processed_opk.sort(key=opk_key)
            return processed_opk

    def extract_standard(self):
        for table_info in self.all_tables:
            table_data = table_info['data']

            table_text = ' '.join([' '.join(row) for row in table_data]).lower()
            has_standards = 'профессиональн' in table_text and 'стандарт' in table_text
            has_code = 'код' in table_text

            if has_standards and has_code:
                return table_data

    def _auto_adjust_columns(self, worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

                if cell.value:
                    lines = str(cell.value).split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    max_length = max(max_length, max_line_length)

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def save_all_to_excel(self, filename=None):
        if filename is None:
            doc_name = self.extract_doc_name()
            filename = f"{doc_name}.xlsx"
        elif not filename.endswith('.xlsx'):
            filename += '.xlsx'

        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 1. Дисциплины специальности
                discipliny = self.extract_discipliny_specialiteta()
                if discipliny:
                    pd.DataFrame(discipliny).to_excel(writer, sheet_name='Дисциплины', index=False, header=False)

                # 2. Специализации
                specializacii = self.extract_specializacii()
                if specializacii and len(specializacii) == 2:
                    numbers_s, names_s = specializacii
                    data = list(zip(numbers_s, names_s))
                    pd.DataFrame(data).to_excel(writer, sheet_name='Специализации', index=False, header=False)

                # 3. Структура и объем
                structure = self.extract_structure_and_volume()
                if structure:
                    pd.DataFrame(structure).to_excel(writer, sheet_name='Структура_и_объем', index=False, header=False)

                # 4. Практика - сохраняем в один лист с двумя колонками
                praktika = self.extract_praktika()
                if praktika and len(praktika) == 2:
                    uch_praktika, pr_praktika = praktika

                    max_len = max(len(uch_praktika), len(pr_praktika))
                    data = []
                    for i in range(max_len):
                        uch = uch_praktika[i] if i < len(uch_praktika) else ''
                        pr = pr_praktika[i] if i < len(pr_praktika) else ''
                        data.append([uch, pr])

                    pd.DataFrame(data).to_excel(writer, sheet_name='Практика', index=False, header=False)

                # 5. Универсальные компетенции
                uk = self.extract_uk()
                if uk:
                    pd.DataFrame(uk).to_excel(writer, sheet_name='Универсальные_компетенции', index=False, header=False)

                # 6. ОПК
                opk = self.extract_opk()
                if opk:
                    pd.DataFrame(opk).to_excel(writer, sheet_name='ОПК', index=False, header=False)

                # 7. Профессиональные стандарты
                standards = self.extract_standard()
                if standards:
                    pd.DataFrame(standards).to_excel(writer, sheet_name='Проф_стандарты', index=False, header=False)


            workbook = load_workbook(filename)
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                self._auto_adjust_columns(worksheet)

            workbook.save(filename)
            print(f"Файл успешно сохранен: {filename}")
            return filename

        except Exception as e:
            print(f"Ошибка при сохранении файла: {e}")
            return None

if __name__ == "__main__":
    url = "https://fgos.ru/fgos/fgos-02-04-01-matematika-i-kompyuternye-nauki-810/"

    parser = Parser(url)

    if parser.soup:

        excel_file = parser.save_all_to_excel()

        print(f"\nФайл сохранен: {excel_file}")